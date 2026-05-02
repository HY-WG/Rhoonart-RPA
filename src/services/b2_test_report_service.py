from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any

import gspread
from gspread.exceptions import WorksheetNotFound
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.core.crawlers.naver_clip_crawler import NaverClipCrawler
from src.core.notifiers.email_notifier import EmailNotifier
from src.core.repositories.supabase_b2_repository import SupabaseB2Repository

KST = pytz.timezone("Asia/Seoul")
REPORT_HEADERS = [
    "video_url",
    "uploaded_at",
    "channel_name",
    "view_count",
    "checked_at",
    "clip_title",
    "work_title",
    "platform",
    "rights_holder_name",
    "identifier",
]


class B2TestReportService:
    def __init__(
        self,
        *,
        repository: SupabaseB2Repository,
        sheets_client: gspread.Client | None = None,
        email_notifier: EmailNotifier | None = None,
        max_clips_per_identifier: int = 2000,
    ) -> None:
        self._repository = repository
        self._sheets_client = sheets_client
        self._email_notifier = email_notifier
        self._max_clips_per_identifier = max_clips_per_identifier

    def seed_requested_rows(
        self,
        *,
        recipient_email: str,
    ) -> dict[str, Any]:
        content = self._repository.upsert_content_catalog_item(
            content_name="현상수배",
            identifier="1UBvb",
            rights_holder_name="이놀미디어",
        )
        holder = self._repository.upsert_rights_holder(
            rights_holder_name="이놀미디어",
            email=recipient_email,
            current_work_title="현상수배",
            naver_report_enabled=True,
            update_cycle="daily 10:00 KST",
        )
        return {"content": content, "rights_holder": holder}

    def collect_enabled_reports(self, *, triggered_by: str = "manual") -> list[dict[str, Any]]:
        self._repository.ensure_daily_clip_reports_table()
        catalog = self._repository.list_enabled_content_catalog()
        holders = {
            row.get("rights_holder_name"): row
            for row in self._repository.list_rights_holders(enabled_only=True, limit=1000)
            if row.get("rights_holder_name")
        }
        contents = [
            (str(row["identifier"]), str(row["content_name"]))
            for row in catalog
            if row.get("identifier") and row.get("content_name")
        ]
        catalog_by_identifier = {
            str(row["identifier"]): row
            for row in catalog
            if row.get("identifier")
        }
        if not contents:
            return []

        checked_at = datetime.now(KST).isoformat()
        run = self._repository.create_daily_report_run(
            checked_at=checked_at,
            triggered_by=triggered_by,
            target_identifier_count=len(contents),
        )
        run_id = str(run["run_id"])
        crawler = NaverClipCrawler(
            contents=contents,
            max_clips=self._max_clips_per_identifier,
            use_parallel=True,
            max_workers=4,
        )
        rows: list[dict[str, Any]] = []
        try:
            for stat in crawler.crawl_stats():
                catalog_row = catalog_by_identifier.get(stat.identifier, {})
                rights_holder_name = catalog_row.get("rights_holder_name")
                holder = holders.get(rights_holder_name, {})
                for clip in stat.clips:
                    rows.append(
                        {
                            "video_url": clip.video_url,
                            "uploaded_at": clip.published_time.date().isoformat()
                            if clip.published_time
                            else None,
                            "channel_name": clip.nickname or clip.profile_id,
                            "view_count": clip.views,
                            "checked_at": checked_at,
                            "clip_title": clip.title,
                            "work_title": stat.content_name,
                            "platform": "naver_clip",
                            "rights_holder_name": rights_holder_name,
                            "identifier": stat.identifier,
                            "content_catalog_id": catalog_row.get("id"),
                            "rights_holder_id": holder.get("id"),
                        }
                    )
            rows = self._dedupe_by_video_url(rows)
            self._repository.insert_daily_clip_reports(run_id=run_id, rows=rows)
            self._repository.finish_daily_report_run(
                run_id=run_id,
                status="success",
                row_count=len(rows),
            )
            self._repository.refresh_yearly_clip_reports(datetime.fromisoformat(checked_at).year)
            return rows
        except Exception as exc:
            self._repository.finish_daily_report_run(
                run_id=run_id,
                status="failed",
                row_count=len(rows),
                error_message=str(exc),
            )
            raise

    def publish_holder_sheets(
        self,
        rows: list[dict[str, Any]],
        *,
        share_with: str | None = None,
    ) -> dict[str, str]:
        if self._sheets_client is None:
            return {}

        urls: dict[str, str] = {}
        holders = {
            row["rights_holder_name"]: row
            for row in self._repository.list_rights_holders(enabled_only=True, limit=1000)
            if row.get("rights_holder_name")
        }
        for holder_name, holder_rows in self._group_by_holder(rows).items():
            holder = holders.get(holder_name, {})
            existing_url = holder.get("looker_spreadsheet_url")
            if existing_url:
                spreadsheet = self._sheets_client.open_by_url(existing_url)
            else:
                title = f"B-2 Naver Clip Test Report - {holder_name}"
                spreadsheet = self._sheets_client.create(title)
                self._repository.update_rights_holder_report_links(
                    rights_holder_name=holder_name,
                    looker_spreadsheet_url=spreadsheet.url,
                )
            worksheet = self._get_or_create_worksheet(spreadsheet, "b2_clip_reports_test")
            worksheet.clear()
            values = [REPORT_HEADERS] + [
                [row.get(header, "") for header in REPORT_HEADERS]
                for row in holder_rows
            ]
            worksheet.update("A1", values, value_input_option="USER_ENTERED")
            worksheet.freeze(rows=1)
            if share_with:
                spreadsheet.share(share_with, perm_type="user", role="reader")
            urls[holder_name] = spreadsheet.url
        return urls

    def build_workbook_bytes(self, rows: list[dict[str, Any]]) -> bytes:
        wb = Workbook()
        first = True
        for holder_name, holder_rows in self._group_by_holder(rows).items():
            ws = wb.active if first else wb.create_sheet()
            first = False
            ws.title = self._safe_sheet_title(holder_name)
            ws.append(REPORT_HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
            for row in holder_rows:
                ws.append([row.get(header, "") for header in REPORT_HEADERS])
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
                ws.column_dimensions[column_cells[0].column_letter].width = width

        if first:
            ws = wb.active
            ws.title = "empty"
            ws.append(REPORT_HEADERS)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def send_email(
        self,
        *,
        recipient: str,
        rows: list[dict[str, Any]],
        holder_sheet_urls: dict[str, str],
        looker_urls: dict[str, str],
    ) -> bool:
        if self._email_notifier is None:
            raise RuntimeError("email notifier is not configured")

        workbook = self.build_workbook_bytes(rows)
        today = datetime.now(KST).strftime("%Y-%m-%d")
        body = self._build_email_body(
            rows=rows,
            holder_sheet_urls=holder_sheet_urls,
            looker_urls=looker_urls,
        )
        return self._email_notifier.send(
            recipient=recipient,
            subject=f"[Rhoonart] B-2 Naver Clip test report {today}",
            message=body,
            html=True,
            attachments=[(f"b2_clip_reports_test_{today}.xlsx", workbook)],
        )

    def get_enabled_looker_urls(self) -> dict[str, str]:
        holders = self._repository.list_rights_holders(enabled_only=True, limit=1000)
        return {
            row["rights_holder_name"]: row.get("looker_studio_url") or ""
            for row in holders
            if row.get("rights_holder_name")
        }

    def _build_email_body(
        self,
        *,
        rows: list[dict[str, Any]],
        holder_sheet_urls: dict[str, str],
        looker_urls: dict[str, str],
    ) -> str:
        total_views = sum(int(row.get("view_count") or 0) for row in rows)
        sections = []
        for holder_name, holder_rows in self._group_by_holder(rows).items():
            sheet_url = holder_sheet_urls.get(holder_name, "")
            looker_url = looker_urls.get(holder_name, "")
            looker_line = (
                f'<li>Looker Studio: <a href="{looker_url}">{looker_url}</a></li>'
                if looker_url
                else '<li>Looker Studio: created Google Sheet is ready for connector setup.</li>'
            )
            sections.append(
                f"""
                <h3>{holder_name}</h3>
                <ul>
                  <li>Rows: {len(holder_rows)}</li>
                  <li>Google Sheet: <a href="{sheet_url}">{sheet_url}</a></li>
                  {looker_line}
                </ul>
                """
            )
        return f"""
        <html><body style="font-family:Arial,sans-serif;color:#111827;">
          <p>B-2 Naver Clip test report has been generated.</p>
          <p>Total rows: <strong>{len(rows)}</strong><br/>
             Total views: <strong>{total_views:,}</strong></p>
          {''.join(sections)}
          <p>The Excel workbook is attached.</p>
        </body></html>
        """

    @staticmethod
    def _group_by_holder(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            grouped[str(row.get("rights_holder_name") or "unknown")].append(row)
        return dict(sorted(grouped.items()))

    @staticmethod
    def _dedupe_by_video_url(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        by_url: dict[str, dict[str, Any]] = {}

        def sort_key(row: dict[str, Any]) -> tuple[int, str, str]:
            return (
                int(row.get("view_count") or 0),
                str(row.get("uploaded_at") or ""),
                str(row.get("checked_at") or ""),
            )

        for row in rows:
            video_url = str(row.get("video_url") or "")
            if not video_url:
                continue
            existing = by_url.get(video_url)
            if existing is None or sort_key(row) > sort_key(existing):
                by_url[video_url] = row
        return list(by_url.values())

    @staticmethod
    def _safe_sheet_title(value: str) -> str:
        for char in ["\\", "/", "?", "*", "[", "]", ":"]:
            value = value.replace(char, " ")
        return (value or "unknown")[:31]

    @staticmethod
    def _get_or_create_worksheet(spreadsheet: gspread.Spreadsheet, title: str) -> gspread.Worksheet:
        try:
            return spreadsheet.worksheet(title)
        except WorksheetNotFound:
            return spreadsheet.add_worksheet(title=title, rows=1000, cols=len(REPORT_HEADERS))
