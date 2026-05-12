"""A-3 monthly Naver Clip settlement workflow.

Modes:
  - confirm: notify Slack with this month's Naver Clip activity list
  - send: email the previous month's Google Sheet link
"""
from __future__ import annotations

import io
from datetime import datetime
from enum import Enum
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pytz

from ..core.interfaces.notifier import INotifier
from ..core.interfaces.repository import ILogRepository, INaverClipRepository
from ..core.logger import CoreLogger
from ..models import NaverClipApplicant

KST = pytz.timezone("Asia/Seoul")
log = CoreLogger(__name__)

TASK_ID = "A-3"
TASK_NAME = "Naver Clip monthly onboarding intake"


class RunMode(str, Enum):
    CONFIRM = "confirm"
    SEND = "send"
    SEND_DUE = "send_due"


def run(
    form_repo: INaverClipRepository,
    log_repo: ILogRepository,
    slack_notifier: INotifier,
    email_notifier: INotifier,
    mode: RunMode,
    manager_email: str,
    target_year: Optional[int] = None,
    target_month: Optional[int] = None,
    slack_channel: str = "",
    holiday_dates: Optional[list[str]] = None,
) -> dict:
    if mode == RunMode.SEND_DUE and not _is_monthly_send_due_today(holiday_dates or []):
        today = datetime.now(KST).date().isoformat()
        return {"mode": "send_due", "action": "skipped_not_due", "date": today}

    year, month = _resolve_target_month(target_year, target_month, mode)
    applicants = form_repo.get_applicants_by_month(year, month)
    sheet_url = ""
    if hasattr(form_repo, "month_sheet_url"):
        sheet_url = form_repo.month_sheet_url(year, month)  # type: ignore[attr-defined]

    if mode == RunMode.CONFIRM:
        return _handle_confirm(applicants, year, month, slack_notifier, slack_channel, sheet_url)
    return _handle_send(applicants, year, month, email_notifier, manager_email, sheet_url)


def _handle_confirm(
    applicants: list[NaverClipApplicant],
    year: int,
    month: int,
    slack_notifier: INotifier,
    slack_channel: str,
    sheet_url: str = "",
) -> dict:
    if not applicants:
        message = (
            f":information_source: *네이버 클립 활동 명단 확인 요청* ({year}-{month:02d})\n"
            "이번 달 신규 정산 신청 데이터가 없습니다.\n"
            f"시트: {sheet_url or '-'}"
        )
        slack_notifier.send(slack_channel, message)
        return {"mode": "confirm", "applicant_count": 0, "action": "no_applicants_notified"}

    lines = [
        f"*[네이버 클립 활동 명단 확인 요청]* {year}-{month:02d}",
        f"총 {len(applicants)}건의 정산 신청 데이터가 있습니다.",
        "변경사항이 없으면 차월 5일에 아래 Google Sheet 링크를 포함한 메일이 자동 발송됩니다.",
        f"시트: {sheet_url or '-'}",
        "",
    ]
    for index, applicant in enumerate(applicants, start=1):
        lines.append(
            f"{index}. {applicant.name} | {applicant.representative_channel_name} | "
            f"{applicant.representative_channel_platform.value} | "
            f"{applicant.naver_clip_profile_name}"
        )

    slack_notifier.send(slack_channel, "\n".join(lines))
    return {"mode": "confirm", "applicant_count": len(applicants), "action": "slack_sent"}


def _handle_send(
    applicants: list[NaverClipApplicant],
    year: int,
    month: int,
    email_notifier: INotifier,
    manager_email: str,
    sheet_url: str = "",
) -> dict:
    if not applicants:
        return {"mode": "send", "applicant_count": 0, "action": "skipped_no_applicants"}

    subject = f"[Rhoonart] {year}-{month:02d} 네이버 클립 수익금 정보"
    body = _build_email_body(len(applicants), year, month, sheet_url)

    success = email_notifier.send(
        recipient=manager_email,
        message=body,
        subject=subject,
        html=True,
    )
    if not success:
        raise RuntimeError(f"failed to send onboarding email to {manager_email}")
    return {
        "mode": "send",
        "applicant_count": len(applicants),
        "action": "email_sent",
        "sheet_url": sheet_url,
    }


def _build_excel(applicants: list[NaverClipApplicant], year: int, month: int) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{year}-{month:02d}"

    headers = [
        "No.",
        "신청일시",
        "이름",
        "전화번호",
        "네이버 ID",
        "네이버 클립 프로필명",
        "네이버 클립 프로필 ID",
        "대표 채널명",
        "대표 채널의 활동 플랫폼",
        "채널 URL",
    ]
    _write_header_row(worksheet, headers)

    for index, applicant in enumerate(applicants, start=1):
        worksheet.append(
            [
                index,
                applicant.submitted_at.strftime("%Y-%m-%d %H:%M"),
                applicant.name,
                applicant.phone_number,
                applicant.naver_id,
                applicant.naver_clip_profile_name,
                applicant.naver_clip_profile_id,
                applicant.representative_channel_name,
                applicant.representative_channel_platform.value,
                applicant.channel_url,
            ]
        )
        if applicant.channel_url:
            worksheet.cell(row=index + 1, column=10).hyperlink = applicant.channel_url
            worksheet.cell(row=index + 1, column=10).font = Font(
                color="0563C1",
                underline="single",
            )

    widths = [6, 18, 14, 18, 18, 24, 22, 22, 26, 40]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_header_row(worksheet, headers: list[str]) -> None:
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    worksheet.append(headers)
    for column_index, _ in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border
    worksheet.row_dimensions[1].height = 22


def _build_email_body(count: int, year: int, month: int, sheet_url: str = "") -> str:
    link_html = (
        f"<p><a href='{sheet_url}' target='_blank' rel='noreferrer'>Google Sheet 열기</a></p>"
        if sheet_url
        else "<p>Google Sheet 링크가 설정되지 않았습니다.</p>"
    )
    return (
        "<html><body style='font-family:sans-serif;color:#333;'>"
        "<p>안녕하세요.</p>"
        f"<p>{year}-{month:02d} 네이버 클립 수익금 정보 시트를 전달드립니다.</p>"
        f"<p>정산 대상 수: <strong>{count}명</strong></p>"
        f"{link_html}"
        "<p style='font-size:12px;color:#888;'>본 메일은 자동 발송되었습니다.</p>"
        "</body></html>"
    )


def _resolve_target_month(
    year: Optional[int],
    month: Optional[int],
    mode: RunMode,
) -> tuple[int, int]:
    if year and month:
        return year, month

    now = datetime.now(KST)
    if mode in {RunMode.SEND, RunMode.SEND_DUE}:
        if now.month == 1:
            return now.year - 1, 12
        return now.year, now.month - 1
    return now.year, now.month


def _is_monthly_send_due_today(holiday_dates: list[str]) -> bool:
    today = datetime.now(KST).date()
    if today.day < 5:
        return False
    holidays = set(holiday_dates)

    cursor = today.replace(day=5)
    while cursor.weekday() >= 5 or cursor.isoformat() in holidays:
        cursor = cursor.fromordinal(cursor.toordinal() + 1)
    return today == cursor
