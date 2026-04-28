from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pytz
from openpyxl import load_workbook

from src.handlers.a3_naver_clip_monthly import RunMode, run
from src.models import NaverClipApplicant, RepresentativeChannelPlatform
from tests.fakes import FakeFormRepo, FakeLogRepo, FakeNotifier

KST = pytz.timezone("Asia/Seoul")


def _build_applicant() -> NaverClipApplicant:
    return NaverClipApplicant(
        applicant_id="a3-test-001",
        name="홍길동",
        phone_number="010-1234-5678",
        naver_id="naver_user_01",
        naver_clip_profile_name="홍길동 클립",
        naver_clip_profile_id="clip-profile-001",
        representative_channel_name="대표 채널 A",
        representative_channel_platform=RepresentativeChannelPlatform.YOUTUBE,
        channel_url="https://example.com/channel-a",
        submitted_at=datetime(2026, 4, 20, 10, 0, tzinfo=KST),
    )


def test_a3_confirm_no_applicants_sends_slack_notice() -> None:
    form_repo = FakeFormRepo([])
    slack_notifier = FakeNotifier(send_result=True)
    email_notifier = FakeNotifier(send_result=True)

    result = run(
        form_repo=form_repo,
        log_repo=FakeLogRepo(),
        slack_notifier=slack_notifier,
        email_notifier=email_notifier,
        mode=RunMode.CONFIRM,
        manager_email="manager@example.com",
        target_year=2026,
        target_month=4,
        slack_channel="#naver-clip",
    )

    assert result == {
        "mode": "confirm",
        "applicant_count": 0,
        "action": "no_applicants_notified",
    }
    assert slack_notifier.sent[0]["recipient"] == "#naver-clip"


def test_a3_send_builds_excel_and_emails_attachment() -> None:
    applicant = _build_applicant()
    form_repo = FakeFormRepo([applicant])
    email_notifier = FakeNotifier(send_result=True)

    result = run(
        form_repo=form_repo,
        log_repo=FakeLogRepo(),
        slack_notifier=FakeNotifier(send_result=True),
        email_notifier=email_notifier,
        mode=RunMode.SEND,
        manager_email="manager@example.com",
        target_year=2026,
        target_month=4,
    )

    assert result["mode"] == "send"
    assert result["applicant_count"] == 1
    assert result["action"] == "email_sent"
    sent = email_notifier.sent[0]
    assert sent["recipient"] == "manager@example.com"
    attachments = sent["kwargs"]["attachments"]
    assert len(attachments) == 1

    filename, payload = attachments[0]
    assert filename.endswith(".xlsx")

    workbook = load_workbook(BytesIO(payload))
    ws = workbook.active
    assert ws["C2"].value == "홍길동"
    assert ws["E2"].value == "naver_user_01"
    assert ws["I2"].value == "유튜브"
    assert ws["J2"].value == "https://example.com/channel-a"
